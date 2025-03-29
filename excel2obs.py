import pandas as pd
import openpyxl
import websocket
import json
from tkinter import ttk, Tk, filedialog, Label, Entry, Button, Frame, Checkbutton, IntVar, OptionMenu, StringVar, messagebox, Scrollbar
from tkinter.ttk import Combobox
import logging
import os
import threading
import time

# 设置日志
logging.basicConfig(level=logging.DEBUG, format='%(asctime)s - %(levelname)s - %(message)s')

# OBS WebSocket 地址和端口
obs_ws_url = "ws://localhost:4444"

class ExcelToOBS:
    def __init__(self, root):
        self.root = root
        self.root.title("Excel2OBS - 更美观版 作者 B站:直播说")
        self.root.geometry("800x600")
        
        # 设置现代主题
        self.style = ttk.Style()
        self.style.theme_use('clam')
        
        # 主容器
        self.main_frame = ttk.Frame(root)
        self.main_frame.pack(fill='both', expand=True, padx=10, pady=10)

        # 文件选择区域
        self.file_frame = ttk.LabelFrame(self.main_frame, text="Excel文件配置")
        self.file_frame.pack(fill='x', pady=5)
        
        ttk.Label(self.file_frame, text="Excel文件:").grid(row=0, column=0, padx=5)
        self.file_entry = ttk.Entry(self.file_frame, width=40)
        self.file_entry.grid(row=0, column=1, padx=5)
        ttk.Button(self.file_frame, text="浏览", command=self.choose_file).grid(row=0, column=2)
        
        ttk.Label(self.file_frame, text="工作表:").grid(row=1, column=0, padx=5)
        self.sheet_combobox = Combobox(self.file_frame, state="readonly")
        self.sheet_combobox.grid(row=1, column=1, padx=5, sticky='ew')

        # 输入配置区域（带滚动条）
        self.config_frame = ttk.LabelFrame(self.main_frame, text="输入配置")
        self.canvas = ttk.Canvas(self.config_frame)
        self.scrollbar = ttk.Scrollbar(self.config_frame, orient="vertical", command=self.canvas.yview)
        self.scrollable_frame = ttk.Frame(self.canvas)

        self.scrollable_frame.bind(
            "<Configure>",
            lambda e: self.canvas.configure(scrollregion=self.canvas.bbox("all"))
        )

        self.canvas.create_window((0, 0), window=self.scrollable_frame, anchor="nw")
        self.canvas.configure(yscrollcommand=self.scrollbar.set)

        self.config_frame.pack(fill='both', expand=True, pady=5)
        self.canvas.pack(side="left", fill="both", expand=True)
        self.scrollbar.pack(side="right", fill="y")

        # 控制按钮区域
        self.control_frame = ttk.Frame(self.main_frame)
        self.control_frame.pack(fill='x', pady=5)
        
        ttk.Button(self.control_frame, text="添加输入项", command=self.add_input).pack(side='left', padx=5)
        ttk.Button(self.control_frame, text="立即更新", command=lambda: self.update_text(check_changes=False)).pack(side='left', padx=5)
        self.status_label = ttk.Label(self.control_frame, text="就绪")
        self.status_label.pack(side='right', padx=5)

        # 初始化变量
        self.file_path = None
        self.inputs = []
        self.previous_values = {}
        self.running = True
        
        self.add_input()
        self.start_update_thread()

    def choose_file(self):
        """选择Excel文件并加载工作表"""
        file_path = filedialog.askopenfilename(filetypes=[("Excel文件", "*.xlsx;*.xlsm")])
        if file_path:
            self.file_entry.delete(0, 'end')
            self.file_entry.insert(0, file_path)
            self.file_path = file_path
            self.load_sheets()
            self.status_label.config(text="文件已加载")

    def load_sheets(self):
        """加载工作表列表"""
        try:
            wb = openpyxl.load_workbook(self.file_path, read_only=True)
            self.sheet_combobox['values'] = wb.sheetnames
            if wb.sheetnames:
                self.sheet_combobox.current(0)
        except Exception as e:
            messagebox.showerror("错误", f"读取工作表失败: {str(e)}")

    def add_input(self):
        """添加新的输入配置项"""
        row_index = len(self.inputs)
        input_frame = ttk.Frame(self.scrollable_frame, relief="groove", borderwidth=1)
        input_frame.pack(fill='x', pady=2, padx=5)

        # 数据类型选择
        data_type_var = StringVar(value="Text")
        ttk.OptionMenu(input_frame, data_type_var, "Text", "Text", "Image").grid(row=0, column=0, padx=2)

        # 源名称
        ttk.Label(input_frame, text="源名称:").grid(row=0, column=1)
        name_entry = ttk.Entry(input_frame, width=15)
        name_entry.grid(row=0, column=2, padx=2)

        # 行列输入
        ttk.Label(input_frame, text="行:").grid(row=0, column=3)
        row_entry = ttk.Entry(input_frame, width=4)
        row_entry.grid(row=0, column=4, padx=2)

        ttk.Label(input_frame, text="列:").grid(row=0, column=5)
        column_entry = ttk.Entry(input_frame, width=4)
        column_entry.grid(row=0, column=6, padx=2)

        # 实时值显示
        value_label = ttk.Label(input_frame, text="N/A", width=10)
        value_label.grid(row=0, column=7, padx=5)

        # 自动更新复选框
        check_var = IntVar()
        ttk.Checkbutton(input_frame, text="自动更新", variable=check_var).grid(row=0, column=8, padx=5)

        # 删除按钮
        ttk.Button(input_frame, text="×", width=2, 
                 command=lambda f=input_frame: self.remove_input(f)).grid(row=0, column=9)

        # 绑定事件
        row_entry.bind("<KeyRelease>", lambda e: self.update_value_label(row_entry, column_entry, value_label))
        column_entry.bind("<KeyRelease>", lambda e: self.update_value_label(row_entry, column_entry, value_label))

        self.inputs.append((input_frame, data_type_var, row_entry, column_entry, name_entry, value_label, check_var))

    def remove_input(self, frame):
        """删除输入配置项"""
        for item in self.inputs:
            if item[0] == frame:
                frame.destroy()
                self.inputs.remove(item)
                break

    # 其他方法保持不变，仅修改GUI相关部分...

    def update_value_label(self, row_entry, column_entry, value_label):
        """更新值标签（添加错误提示）"""
        try:
            # ...原有逻辑...
            value_label.config(text=str(value), foreground='green')
        except Exception as e:
            value_label.config(text="错误", foreground='red')
            self.status_label.config(text=f"错误: {str(e)}")

    def update_text(self, check_changes=False):
        """更新文本（添加状态提示）"""
        try:
            # ...原有逻辑...
            self.status_label.config(text="更新成功", foreground='green')
        except Exception as e:
            self.status_label.config(text=f"更新失败: {str(e)}", foreground='red')
            logging.error(str(e))

    # 其他方法保持不变...

root = Tk()
app = ExcelToOBS(root)
root.protocol("WM_DELETE_WINDOW", app.stop)
root.mainloop()